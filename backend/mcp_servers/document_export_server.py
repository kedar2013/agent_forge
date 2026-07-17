"""A real MCP server for exporting an agent's analysis as a downloadable
PDF or Excel file — pure local computation (reportlab / openpyxl), no
external API. Runs over stdio, spawned as a subprocess by Agent Forge's
mcp_tool (StdioConnectionParams), following the same pattern as the other
mcp_servers/*.py files.

Files are written to backend/generated_files/, which app/main.py mounts at
/generated-files — the same pattern app/tool_registry/image_gen_tool.py
uses for generated images. Returns an ABSOLUTE URL (via BACKEND_PUBLIC_URL)
rather than a relative one, since the frontend and backend run on different
origins/ports in dev — a relative link would resolve against the frontend's
origin and 404.

Run standalone for a smoke test:
    python mcp_servers/document_export_server.py
"""

import os
import re
import uuid
from typing import Any

from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from _signed_urls import sign_filename
from _watermark import render_excel_watermark_image, watermark_pdf_page

mcp = FastMCP("document-export")


def _excel_cell_value(value: Any) -> Any:
    """Money/decimal values commonly arrive as JSON *strings* (e.g. from
    data_query_tool, which stringifies Decimals to avoid float precision
    loss) — written as-is, openpyxl stores them as Excel *text*, not
    numbers, losing right-alignment/SUM()/number-formatting in the sheet.
    Converts anything that parses cleanly as a number; leaves genuine text
    (names, ids, dates already formatted as strings) untouched."""
    if not isinstance(value, str):
        return value
    try:
        return int(value) if value.lstrip("-").isdigit() else float(value)
    except (ValueError, AttributeError):
        return value

# Matches the directory app/main.py mounts at /generated-files, regardless
# of the process's current working directory.
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "generated_files")
PUBLIC_BASE_URL = os.environ.get("BACKEND_PUBLIC_URL", "http://127.0.0.1:8000")

_BOLD_RE = re.compile(r"\*\*(.+?)\*\*")
_ITALIC_RE = re.compile(r"(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)")
_INLINE_CODE_RE = re.compile(r"`(.+?)`")


def _inline_markdown_to_reportlab(text: str) -> str:
    """Converts a line's **bold**/*italic*/`code` spans to reportlab's
    limited inline-tag markup (Paragraph supports a small HTML-like subset)."""
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    text = _BOLD_RE.sub(r"<b>\1</b>", text)
    text = _ITALIC_RE.sub(r"<i>\1</i>", text)
    text = _INLINE_CODE_RE.sub(r"<font face='Courier'>\1</font>", text)
    return text


def _markdown_to_flowables(markdown_content: str) -> list:
    """Best-effort conversion of the same GFM markdown an agent already
    writes in chat (headings, bold, bullet lists, tables) into reportlab
    flowables — not a full CommonMark parser, but handles what these agents
    actually produce."""
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], spaceAfter=10)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], spaceAfter=8)
    h3 = ParagraphStyle("H3", parent=styles["Heading3"], spaceAfter=6)
    body = ParagraphStyle("Body", parent=styles["BodyText"], spaceAfter=6, leading=15)
    bullet = ParagraphStyle("Bullet", parent=body, leftIndent=16, bulletIndent=4)

    flowables: list = []
    lines = markdown_content.replace("\r\n", "\n").split("\n")
    table_buffer: list[list[str]] = []

    def _flush_table() -> None:
        if not table_buffer:
            return
        rows = [[_inline_markdown_to_reportlab(c) for c in row] for row in table_buffer]
        rows = [[Paragraph(cell, body) for cell in row] for row in rows]
        t = Table(rows, hAlign="LEFT")
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#5b3fe6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]
            )
        )
        flowables.append(t)
        flowables.append(Spacer(1, 10))
        table_buffer.clear()

    for raw_line in lines:
        line = raw_line.rstrip()
        stripped = line.strip()

        if stripped.startswith("|") and stripped.endswith("|"):
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            if all(re.fullmatch(r"[:\-\s]+", c) for c in cells):
                continue  # markdown table separator row (---|---)
            table_buffer.append(cells)
            continue
        _flush_table()

        if not stripped:
            flowables.append(Spacer(1, 6))
        elif stripped.startswith("### "):
            flowables.append(Paragraph(_inline_markdown_to_reportlab(stripped[4:]), h3))
        elif stripped.startswith("## "):
            flowables.append(Paragraph(_inline_markdown_to_reportlab(stripped[3:]), h2))
        elif stripped.startswith("# "):
            flowables.append(Paragraph(_inline_markdown_to_reportlab(stripped[2:]), h1))
        elif stripped in ("---", "***", "___"):
            flowables.append(Spacer(1, 12))
        elif stripped.startswith(("- ", "* ")):
            flowables.append(Paragraph("&bull;&nbsp;&nbsp;" + _inline_markdown_to_reportlab(stripped[2:]), bullet))
        else:
            flowables.append(Paragraph(_inline_markdown_to_reportlab(stripped), body))

    _flush_table()
    return flowables


def _safe_filename(title: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "-", title).strip("-").lower() or "document"
    return f"{slug[:60]}-{uuid.uuid4().hex[:8]}"


@mcp.tool()
async def export_to_pdf(title: str, markdown_content: str) -> str:
    """Generate a formatted PDF from analysis you've already written, and
    get back a download link. Use this when the user asks to save, export,
    download, or get a PDF/report/document of something you just told them.

    Args:
        title: A short title for the document (used as filename + PDF header).
        markdown_content: The full write-up in the same markdown you'd
            normally reply with — headings (# ## ###), **bold**, *italic*,
            bullet lists (- item), and GFM tables (| a | b |) all render.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"{_safe_filename(title)}.pdf"
    path = os.path.join(OUTPUT_DIR, filename)

    doc = SimpleDocTemplate(
        path, pagesize=LETTER, topMargin=0.75 * inch, bottomMargin=0.75 * inch,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch, title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("DocTitle", parent=styles["Title"], textColor=colors.HexColor("#372683"))
    story: list = [Paragraph(_inline_markdown_to_reportlab(title), title_style), Spacer(1, 14)]
    story.extend(_markdown_to_flowables(markdown_content))

    def _stamp(canvas, _doc) -> None:
        watermark_pdf_page(canvas, LETTER[0], LETTER[1])

    try:
        doc.build(story, onFirstPage=_stamp, onLaterPages=_stamp)
    except Exception as exc:  # noqa: BLE001 — surfaced to the agent, not the caller
        return f"Couldn't generate the PDF: {exc}"

    token = sign_filename("generated-files", filename)
    return f"PDF generated: {PUBLIC_BASE_URL}/generated-files/{filename}?token={token}"


@mcp.tool()
async def export_to_excel(title: str, sheet_name: str, rows: list[dict[str, Any]]) -> str:
    """Generate a formatted Excel spreadsheet from tabular data you've
    already gathered, and get back a download link. Use this when the user
    asks to export, download, or get a spreadsheet/Excel/CSV of data —
    e.g. a fund comparison table or sector performance breakdown.

    Args:
        title: A short title (used as filename).
        sheet_name: Name for the worksheet tab (max 31 characters).
        rows: The data rows, each a dict of column_name -> value. All rows
            should use the same keys so columns line up.
    """
    if not rows:
        return "No data rows were provided to export."

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"{_safe_filename(title)}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)

    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet1")[:31]

    header_fill = PatternFill(start_color="5B3FE6", end_color="5B3FE6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_excel_cell_value(row.get(col_name)))

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(r.get(col_name, ""))) for r in rows])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

    ws.freeze_panes = "A2"

    width_px = sum(ws.column_dimensions[get_column_letter(c)].width for c in range(1, len(columns) + 1)) * 7
    height_px = (len(rows) + 1) * 20
    watermark = XLImage(render_excel_watermark_image(int(width_px), int(height_px)))
    ws.add_image(watermark, "A1")

    try:
        wb.save(path)
    except Exception as exc:  # noqa: BLE001
        return f"Couldn't generate the spreadsheet: {exc}"

    token = sign_filename("generated-files", filename)
    return f"Excel file generated: {PUBLIC_BASE_URL}/generated-files/{filename}?token={token}"


if __name__ == "__main__":
    mcp.run(transport="stdio")
